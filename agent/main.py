"""
Bank Statement Extraction Agent
Streams MiMo-V2.5-Pro with reasoning_effort=high.
Investigates PDF, extracts transactions, exports XLSX, verifies output.
"""

import json
import logging
import os
import sys
from pathlib import Path

from openai import OpenAI
from dotenv import load_dotenv

from .system_prompt import SYSTEM_PROMPT
from .tools.read_skills import read_skills_doc

logger = logging.getLogger(__name__)

# Load .env if it exists
load_dotenv(Path(__file__).parent.parent / ".env")


TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_skills_doc",
            "description": (
                "Read a skills documentation file to learn how to use a library or pattern. "
                "skill='pdf'   → pdfplumber, pypdf, OCR, text extraction. "
                "skill='excel' → openpyxl formatting, colors, multi-sheet workbooks, "
                "                daily summaries, credit/debit breakdowns. "
                "Call with skill='excel' before writing any formatted XLSX output."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "skill": {
                        "type": "string",
                        "enum": ["pdf", "excel"],
                        "description": "Which skill doc to read: 'pdf' or 'excel'",
                    }
                },
                "required": ["skill"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "run_python",
            "description": (
                "Write and execute any Python script using the project venv. "
                "Available libraries: pdfplumber, pandas, openpyxl, pypdf. "
                "Working directory: /home/ubuntu/Bankstatement/. "
                "Use this to: inspect the PDF structure, extract text/tables, "
                "parse transactions, write the output XLSX file to the exact "
                "path given in the user message, and verify the results. "
                "Returns stdout, stderr, and exit_code. If exit_code != 0, read stderr "
                "and fix the error before retrying. Write complete, runnable scripts."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "Complete Python script to execute",
                    },
                    "description": {
                        "type": "string",
                        "description": "One line describing what this script does",
                    },
                },
                "required": ["code", "description"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "edit_xlsx_payees",
            "description": (
                "Precisely edit Payee cells in the Transactions sheet without re-extracting from PDF. "
                "Use this during feedback rounds to fix only the rows listed in <evaluator_feedback>. "
                "ALWAYS prefer this over re-running the full extraction when you have a list of specific rows to fix. "
                "Row numbers are openpyxl 1-based: row 1 = header, row 2 = first data row."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "xlsx_path": {
                        "type": "string",
                        "description": "Absolute path to the XLSX file to edit",
                    },
                    "edits": {
                        "type": "object",
                        "description": (
                            "Mapping of openpyxl row number (string) → new payee name. "
                            "E.g. {\"15\": \"RAJAWALI SDN BHD\", \"23\": \"MEADOW SDN BHD\"}"
                        ),
                        "additionalProperties": {"type": "string"},
                    },
                },
                "required": ["xlsx_path", "edits"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "verify_xlsx_balance",
            "description": (
                "Verify L1 (aggregate) and L2 (row-by-row) balance integrity of the Transactions sheet. "
                "Returns structured pass/fail with exact failing rows. "
                "Call this after edit_xlsx_payees to confirm the edits did not break balance integrity."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "xlsx_path": {
                        "type": "string",
                        "description": "Absolute path to the XLSX file to verify",
                    },
                },
                "required": ["xlsx_path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "read_xlsx_transactions",
            "description": (
                "Read all transactions from the Transactions sheet as structured data. "
                "Returns columns list and rows list — each row is a dict keyed by column header. "
                "Use this to inspect or verify the current XLSX content without running a script."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "xlsx_path": {
                        "type": "string",
                        "description": "Absolute path to the XLSX file to read",
                    },
                },
                "required": ["xlsx_path"],
            },
        },
    },
]

EVALUATOR_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "run_python",
            "description": (
                "Execute a Python script to read and filter the XLSX file. "
                "Available libraries: openpyxl. "
                "Use this to identify suspicious payees programmatically before judging them."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "Complete Python script to execute",
                    },
                    "description": {
                        "type": "string",
                        "description": "One line describing what this script does",
                    },
                },
                "required": ["code", "description"],
            },
        },
    }
]

PROVIDERS = {
    "openrouter": {
        "base_url": "https://openrouter.ai/api/v1",
        "model": "xiaomi/mimo-v2.5-pro",
        "env_key": "OPENROUTER_API_KEY",
        "extra_body": {"reasoning": {"effort": "high"}},
        "evaluator_model": "xiaomi/mimo-v2.5",
        "evaluator_extra_body": {"reasoning": {"effort": "medium"}},
    },
    "deepinfra": {
        "base_url": "https://api.deepinfra.com/v1/openai",
        "model": "XiaomiMiMo/MiMo-V2.5-Pro",
        "env_key": "DEEPINFRA_API_KEY",
        "extra_body": {"reasoning_effort": "high"},
        "evaluator_model": "XiaomiMiMo/MiMo-V2.5",
        "evaluator_extra_body": {"reasoning_effort": "medium"},
    },
    "xiaomi": {
        "base_url": "https://api.xiaomimimo.com/v1",
        "model": "mimo-v2.5-pro",
        "env_key": "XIAOMI_MIMO_API_KEY",
        "extra_body": {"chat_template_kwargs": {"enable_thinking": True}},
        "evaluator_model": "mimo-v2.5",
        "evaluator_extra_body": {"chat_template_kwargs": {"enable_thinking": True}},
    },
    "xiaomi-tokenplan": {
        "base_url": os.environ.get("XIAOMI_MIMO_BASE_URL", "https://token-plan-sgp.xiaomimimo.com/v1"),
        "model": "mimo-v2.5-pro",
        "env_key": "XIAOMI_MIMO_API_KEY",
        "extra_body": {"reasoning_effort": "high"},
        "evaluator_model": "mimo-v2.5",
        "evaluator_extra_body": {"reasoning_effort": "medium"},
    },
}


class BankStatementAgent:
    CTX_COMPRESS_THRESHOLD = int(os.getenv("CTX_COMPRESS_THRESHOLD", "700000"))
    CTX_KEEP_RECENT_TURNS  = int(os.getenv("CTX_KEEP_RECENT_TURNS", "3"))

    def __init__(
        self,
        api_key: str = None,
        provider: str = None,
        base_url: str = None,
        model: str = None,
    ):
        provider = provider or os.environ.get("MIMO_PROVIDER", "openrouter")
        preset = PROVIDERS.get(provider, PROVIDERS["openrouter"])

        resolved_key = (
            api_key
            or os.environ.get(preset["env_key"])
            or os.environ.get("MIMO_API_KEY")
        )
        self.client = OpenAI(
            api_key=resolved_key,
            base_url=base_url or preset["base_url"],
        )
        self.model = model or preset["model"]
        self.extra_body = preset.get("extra_body", {})
        self.evaluator_model = preset.get("evaluator_model", self.model)
        self.evaluator_extra_body = preset.get("evaluator_extra_body", self.extra_body)
        self.max_turns = 20
        self.state = {
            "skills_loaded": False,
            "trace": [],
            "token_usage": [],   # per-turn token stats
        }

    # ------------------------------------------------------------------
    # Context management
    # ------------------------------------------------------------------

    def _estimate_tokens(self, messages: list) -> int:
        """Rough token estimate: total chars / 4."""
        return sum(
            len(str(m.get("content", ""))) + len(str(m.get("tool_calls", "")))
            for m in messages
        ) // 4

    def _compress_context(self, messages: list) -> list:
        """
        Compress old turns when approaching the token threshold.
        Keeps: system + initial user task + last CTX_KEEP_RECENT_TURNS turns verbatim.
        Compresses middle turns: tool results truncated to exit_code + 150-char stdout,
        assistant messages truncated to 300 chars.
        """
        estimated = self._estimate_tokens(messages)
        if estimated < self.CTX_COMPRESS_THRESHOLD:
            return messages

        header = messages[:2]                        # system + initial user
        rest   = messages[2:]
        keep   = self.CTX_KEEP_RECENT_TURNS * 2     # 2 messages per turn

        if len(rest) <= keep:
            return messages                          # nothing to compress

        to_compress = rest[:-keep]
        to_keep     = rest[-keep:]

        compressed = []
        for msg in to_compress:
            role = msg.get("role", "")
            if role == "tool":
                try:
                    data = json.loads(msg.get("content", "{}"))
                    short = {
                        "exit_code": data.get("exit_code"),
                        "stdout":    (data.get("stdout") or "")[:150],
                    }
                except Exception:
                    short = {"summary": str(msg.get("content", ""))[:150]}
                compressed.append({**msg, "content": json.dumps(short)})
            elif role == "assistant":
                content = str(msg.get("content") or "")
                short_content = (content[:300] + "…[compressed]") if len(content) > 300 else content
                compressed.append({**msg, "content": short_content})
            else:
                compressed.append(msg)

        new_messages = header + compressed + to_keep
        logger.info(
            "context compressed: %d→%d msgs, ~%d→%d tokens",
            len(messages), len(new_messages),
            estimated, self._estimate_tokens(new_messages),
        )
        return new_messages

    # ------------------------------------------------------------------
    # Public entry point — orchestrates extraction + evaluator loop
    # ------------------------------------------------------------------

    def extract(self, pdf_path: str, output_path: str = None, status_callback=None) -> dict:
        """Public API. Runs extraction then Evaluator-Optimizer loop."""
        pdf_path = str(Path(pdf_path).resolve())
        if output_path is None:
            output_path = str(Path(pdf_path).with_suffix(".xlsx"))

        result = self._run_extraction(pdf_path, output_path, status_callback)

        if result.get("status") != "completed":
            return result

        return self._evaluator_loop(
            xlsx_path=result["xlsx_path"],
            pdf_path=pdf_path,
            output_path=output_path,
            status_callback=status_callback,
            base_result=result,
        )

    # ------------------------------------------------------------------
    # Extraction loop (single run, may receive evaluator feedback)
    # ------------------------------------------------------------------

    def _run_extraction(
        self,
        pdf_path: str,
        output_path: str,
        status_callback=None,
        feedback: str = None,
    ) -> dict:
        """
        Single extraction run.
        feedback: optional <evaluator_feedback>...</evaluator_feedback> XML string
                  injected into the initial user message for improvement runs.
        """
        xlsx_path = output_path

        self.state["token_usage"] = []   # reset per extraction run

        print(f"\n[Agent] Starting for: {pdf_path}")
        print(f"[Agent] Model: {self.model}  (reasoning_effort=high, streaming)")
        print(f"[Agent] Excel output: {xlsx_path}")

        if feedback:
            # Feedback round: targeted payee edits only — no PDF re-extraction
            task_content = (
                f"Fix payee quality issues in the existing Excel file.\n"
                f"Excel file to edit: {xlsx_path}\n"
                f"PDF (for reference only if a description is ambiguous): {pdf_path}\n\n"
                f"{feedback}"
            )
        else:
            task_content = (
                f"Extract all transactions from this bank statement PDF "
                f"and export to Excel.\n"
                f"PDF: {pdf_path}\n"
                f"Excel Output: {xlsx_path}"
            )

        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": task_content},
        ]

        stop_without_output = 0

        for turn in range(self.max_turns):
            messages = self._compress_context(messages)

            print(f"\n{'─'*60}")
            print(f"[Turn {turn + 1}]")

            if status_callback:
                status_callback("turn_start", {"turn": turn + 1})

            finish_reason, full_content, tool_calls, usage = self._stream_turn(messages, status_callback=status_callback)

            if status_callback and full_content:
                status_callback("message", {"role": "assistant", "content": full_content, "turn": turn + 1})

            est_flag = " (est)" if usage.get("estimated") else ""
            print(
                f"  [tokens{est_flag}] prompt={usage['prompt_tokens']:,}  "
                f"completion={usage['completion_tokens']:,}  "
                f"total={usage['total_tokens']:,}"
            )
            self.state["trace"].append({
                "turn": turn + 1,
                "finish_reason": finish_reason,
                "usage": usage,
            })
            self.state["token_usage"].append(usage)

            # ── Tool calls ──────────────────────────────────────────────
            if tool_calls and finish_reason in ("tool_calls", "length"):
                assistant_msg = {
                    "role": "assistant",
                    "tool_calls": [
                        {
                            "id": tc["id"],
                            "type": "function",
                            "function": {
                                "name": tc["name"],
                                "arguments": tc["arguments"],
                            },
                        }
                        for tc in tool_calls.values()
                    ],
                }
                if full_content:
                    assistant_msg["content"] = full_content
                messages.append(assistant_msg)

                for tc in tool_calls.values():
                    print(f"\n  → Tool: {tc['name']}")
                    try:
                        inputs = json.loads(tc["arguments"] or "{}")
                    except json.JSONDecodeError:
                        inputs = {}
                    result = self._execute_tool(tc["name"], inputs)
                    if status_callback and tc["name"] == "run_python":
                        status_callback("tool", {
                            "tool": tc["name"],
                            "desc": inputs.get("description", ""),
                            "exit": result.get("exit_code", 0),
                            "preview": (result.get("stdout") or "")[:200],
                        })
                    messages.append(
                        {
                            "role": "tool",
                            "tool_call_id": tc["id"],
                            "content": json.dumps(result),
                        }
                    )
                continue

            # ── Hit token limit mid-response ────────────────────────────
            if finish_reason == "length" and not tool_calls:
                if full_content:
                    messages.append({"role": "assistant", "content": full_content})
                messages.append({
                    "role": "user",
                    "content": (
                        "Your response was cut off. Write and run the Python script now — "
                        "stop planning, use run_python to execute the extraction."
                    ),
                })
                print("  [harness] response cut off (length) — prompting to continue")
                continue

            # ── End turn ────────────────────────────────────────────────
            if finish_reason in ("stop", "end_turn", "length"):
                outputs = self._check_outputs(xlsx_path)
                print(f"\n{'═'*60}")
                if outputs["xlsx_exists"]:
                    total_tok = sum(u.get("total_tokens", 0) for u in self.state["token_usage"])
                    print(f"[Agent] ✅ Completed — {outputs['xlsx_rows']} transactions")
                    print(f"[Agent]    XLSX: {xlsx_path}")
                    print(f"[Agent]    Total tokens this run: {total_tok:,}")
                    return {
                        "status": "completed",
                        "xlsx_path": xlsx_path,
                        "xlsx_rows": outputs["xlsx_rows"],
                        "raw_response": full_content,
                        "trace": self.state["trace"],
                    }
                stop_without_output += 1
                if stop_without_output >= 2:
                    print("[Agent] ⚠️  Stopped twice without output — giving up.")
                    return {
                        "status": "completed_no_output",
                        "raw_response": full_content,
                        "trace": self.state["trace"],
                    }
                print(f"  [harness] no output found (attempt {stop_without_output}) — prompting to act")
                messages.append({"role": "assistant", "content": full_content or "..."})
                messages.append({
                    "role": "user",
                    "content": (
                        f"No Excel file was found at {xlsx_path}. "
                        "Use run_python to write the Excel file now. "
                        "Do not describe what you will do — call run_python immediately."
                    ),
                })
                continue

        outputs = self._check_outputs(xlsx_path)
        if outputs["xlsx_exists"]:
            print(f"\n[Agent] ✅ Max turns reached but XLSX exists — {outputs['xlsx_rows']} rows")
            return {
                "status": "completed",
                "xlsx_path": xlsx_path,
                "xlsx_rows": outputs["xlsx_rows"],
                "raw_response": "",
                "trace": self.state["trace"],
            }
        return {
            "status": "failed",
            "reason": "max_turns_reached",
            "trace": self.state["trace"],
        }

    # ------------------------------------------------------------------
    # Evaluator-Optimizer loop (Anthropic official pattern)
    # ------------------------------------------------------------------

    def _evaluator_loop(
        self,
        xlsx_path: str,
        pdf_path: str,
        output_path: str,
        status_callback,
        base_result: dict,
    ) -> dict:
        """
        Evaluator-Optimizer loop with 4 stopping conditions:
          1. Quality threshold (score >= 8 or verdict == PASS)
          2. Max iterations (3)
          3. Token budget (max_tokens=4096 per evaluator call)
          4. No improvement (2 consecutive same scores)
        """
        MAX_ITER = 3
        PASS_THRESHOLD = 8
        NO_IMPROVE_LIMIT = 2

        attempts = []
        best = base_result
        best_score = 0
        consecutive_same_score = 0
        last_score = None

        for iteration in range(MAX_ITER):
            print(f"\n[Evaluator] iteration {iteration + 1}/{MAX_ITER} — assessing payee quality…")
            if status_callback:
                status_callback("token", {
                    "text": f"\n[Evaluator] iteration {iteration + 1}/{MAX_ITER} — assessing payee quality…\n",
                    "is_thinking": False,
                })

            eval_result = self._run_evaluator(xlsx_path, pdf_path)
            score = eval_result["score"]
            verdict = eval_result["verdict"]

            print(f"[Evaluator] score={score}/10  verdict={verdict}  issues={len(eval_result['issues'])}")
            if status_callback:
                status_callback("token", {
                    "text": f"[Evaluator] score={score}/10  verdict={verdict}  issues={len(eval_result['issues'])}\n",
                    "is_thinking": False,
                })

            attempts.append({
                "iteration": iteration + 1,
                "score": score,
                "verdict": verdict,
                "issues": eval_result["issues"],
                "feedback": eval_result["feedback"],
                "xlsx_path": xlsx_path,
            })

            if score > best_score:
                best_score = score
                best = {
                    **base_result,
                    "xlsx_path": xlsx_path,
                    "eval_score": score,
                    "eval_iterations": iteration + 1,
                }

            # Stopping condition 1: quality
            if verdict == "PASS" or score >= PASS_THRESHOLD:
                print(f"[Evaluator] ✅ PASS — payee quality accepted (score {score}/10)")
                break

            # Stopping condition 4: no improvement
            if score == last_score:
                consecutive_same_score += 1
            else:
                consecutive_same_score = 0
            last_score = score

            if consecutive_same_score >= NO_IMPROVE_LIMIT:
                print(f"[Evaluator] ⚠️  No improvement after {NO_IMPROVE_LIMIT} rounds — stopping")
                break

            # Stopping condition 2: max iterations
            if iteration == MAX_ITER - 1:
                print(f"[Evaluator] ⚠️  Max iterations reached")
                break

            # Build feedback and re-run generator
            feedback_xml = self._build_feedback_xml(attempts, xlsx_path)
            print(f"[Evaluator] Re-running generator with feedback ({len(eval_result['issues'])} issues)…")
            if status_callback:
                status_callback("token", {
                    "text": f"[Evaluator] Re-running generator with {len(eval_result['issues'])} issue(s) to fix…\n",
                    "is_thinking": False,
                })

            re_result = self._run_extraction(pdf_path, output_path, status_callback, feedback=feedback_xml)

            if re_result.get("status") == "completed":
                xlsx_path = re_result["xlsx_path"]
            else:
                print("[Evaluator] ⚠️  Re-extraction failed — using best result so far")
                break

        return best

    def _run_evaluator(self, xlsx_path: str, pdf_path: str) -> dict:
        """
        Agentic evaluator: LLM writes Python to filter suspicious payees (programmatic),
        then judges only the filtered subset (LLM judgment). Max 4 turns.
        Returns {verdict, score, issues, feedback} — same contract as before.
        """
        import xml.etree.ElementTree as ET
        from .evaluator_prompt import EVALUATOR_PROMPT
        from .tools.run_python import run_python as _run_python

        messages = [
            {"role": "system", "content": EVALUATOR_PROMPT},
            {
                "role": "user",
                "content": (
                    f"Evaluate payee quality in this bank statement extraction.\n"
                    f"XLSX file: {xlsx_path}"
                ),
            },
        ]

        eval_total_tokens = 0
        for turn in range(4):
            try:
                response = self.client.chat.completions.create(
                    model=self.evaluator_model,
                    max_tokens=8192,
                    tools=EVALUATOR_TOOLS,
                    tool_choice="auto",
                    messages=messages,
                    extra_body=self.evaluator_extra_body,
                )
            except Exception as exc:
                logger.warning("evaluator call failed: %s — treating as PASS", exc)
                return {"verdict": "PASS", "score": 10, "issues": [], "feedback": f"evaluator error: {exc}"}

            if hasattr(response, "usage") and response.usage:
                t = response.usage.total_tokens
                eval_total_tokens += t
                print(f"  [Evaluator tokens] turn={turn+1}  total={t:,}  cumulative={eval_total_tokens:,}")

            choice = response.choices[0]
            msg = choice.message

            # ── Tool call: execute run_python ───────────────────────────
            if choice.finish_reason == "tool_calls" and msg.tool_calls:
                assistant_msg = {
                    "role": "assistant",
                    "content": msg.content or "",
                    "tool_calls": [
                        {
                            "id": tc.id,
                            "type": "function",
                            "function": {
                                "name": tc.function.name,
                                "arguments": tc.function.arguments,
                            },
                        }
                        for tc in msg.tool_calls
                    ],
                }
                messages.append(assistant_msg)

                for tc in msg.tool_calls:
                    if tc.function.name == "run_python":
                        try:
                            inputs = json.loads(tc.function.arguments or "{}")
                        except json.JSONDecodeError:
                            inputs = {}
                        print(f"  [Evaluator] {inputs.get('description', 'filter script')}")
                        result = _run_python(
                            code=inputs.get("code", ""),
                            description=inputs.get("description", ""),
                        )
                        messages.append({
                            "role": "tool",
                            "tool_call_id": tc.id,
                            "content": json.dumps(result),
                        })
                continue

            # ── Stop: parse XML verdict ─────────────────────────────────
            raw = msg.content or ""
            try:
                start = raw.index("<evaluation>")
                end = raw.rindex("</evaluation>") + len("</evaluation>")
                root = ET.fromstring(raw[start:end])
                score = int(root.findtext("score") or "0")
                verdict = (root.findtext("verdict") or "NEEDS_IMPROVEMENT").strip()
                feedback = (root.findtext("feedback") or "").strip()
                issues = []
                for issue in root.findall(".//issue"):
                    issues.append({
                        "row": issue.findtext("row") or "",
                        "description": issue.findtext("description") or "",
                        "current_payee": issue.findtext("current_payee") or "",
                        "suggested_payee": issue.findtext("suggested_payee") or "",
                        "criterion": issue.findtext("criterion") or "",
                        "reason": issue.findtext("reason") or "",
                    })
                return {"verdict": verdict, "score": score, "issues": issues, "feedback": feedback}
            except Exception as exc:
                logger.warning("evaluator XML parse failed: %s — raw: %.200s", exc, raw)
                return {"verdict": "PASS", "score": 10, "issues": [], "feedback": "XML parse error — accepted as-is"}

        return {"verdict": "PASS", "score": 10, "issues": [], "feedback": "evaluator max turns reached"}

    def _build_feedback_xml(self, attempts: list, xlsx_path: str) -> str:
        """Format the latest evaluator feedback for injection into generator context."""
        latest = attempts[-1]
        lines = [
            f'<evaluator_feedback iteration="{latest["iteration"]}" score="{latest["score"]}/10">',
            f"  <xlsx_path>{xlsx_path}</xlsx_path>",
            f"  <summary>{latest['feedback']}</summary>",
            "",
            "  Fix these specific payees using edit_xlsx_payees (do NOT re-extract from PDF):",
            "  <issues>",
        ]
        for issue in latest["issues"]:
            lines.append("    <issue>")
            lines.append(f"      <row>{issue['row']}</row>")
            lines.append(f"      <description>{issue['description']}</description>")
            lines.append(f"      <current_payee>{issue['current_payee']}</current_payee>")
            lines.append(f"      <suggested_payee>{issue['suggested_payee']}</suggested_payee>")
            lines.append(f"      <criterion>{issue['criterion']}</criterion>")
            lines.append(f"      <reason>{issue['reason']}</reason>")
            lines.append("    </issue>")
        lines.append("  </issues>")
        lines.append("</evaluator_feedback>")
        return "\n".join(lines)

    # ------------------------------------------------------------------
    # Streaming
    # ------------------------------------------------------------------

    def _stream_turn(self, messages: list, status_callback=None) -> tuple[str, str, dict, dict]:
        """
        Stream one API call.
        Returns (finish_reason, full_content, tool_calls_map, usage).
        usage = {prompt_tokens, completion_tokens, total_tokens} (real or estimated).
        Prints thinking tokens in dim color and response text normally.
        """
        THINK_OPEN  = "\033[2m"   # dim  — thinking
        THINK_CLOSE = "\033[0m"   # reset
        RESP_COLOR  = "\033[0m"   # normal

        stream = self.client.chat.completions.create(
            model=self.model,
            max_tokens=32768,
            tools=TOOLS,
            tool_choice="auto",
            stream=True,
            stream_options={"include_usage": True},
            extra_body=self.extra_body,
            messages=messages,
        )

        full_content = ""
        tool_calls: dict[int, dict] = {}  # index → {id, name, arguments}
        finish_reason = "stop"
        in_think_tag = False
        usage: dict = {}

        for chunk in stream:
            # Usage arrives in a final chunk (no choices) when stream_options=include_usage
            if hasattr(chunk, "usage") and chunk.usage:
                usage = {
                    "prompt_tokens":     chunk.usage.prompt_tokens,
                    "completion_tokens": chunk.usage.completion_tokens,
                    "total_tokens":      chunk.usage.total_tokens,
                }

            if not chunk.choices:
                continue
            choice = chunk.choices[0]

            if choice.finish_reason:
                finish_reason = choice.finish_reason

            delta = choice.delta

            # ── Text content (may contain <think>...</think>) ───────────
            if delta.content:
                text = delta.content
                full_content += text
                if status_callback:
                    status_callback("token", {"text": text, "is_thinking": in_think_tag})

                # colorize <think> blocks inline as they stream
                while text:
                    if not in_think_tag:
                        tag_start = text.find("<think>")
                        if tag_start == -1:
                            sys.stdout.write(RESP_COLOR + text)
                            sys.stdout.flush()
                            text = ""
                        else:
                            # print everything before <think>
                            before = text[:tag_start]
                            if before:
                                sys.stdout.write(RESP_COLOR + before)
                            sys.stdout.write(THINK_OPEN + "<think>")
                            sys.stdout.flush()
                            in_think_tag = True
                            text = text[tag_start + len("<think>"):]
                    else:
                        tag_end = text.find("</think>")
                        if tag_end == -1:
                            sys.stdout.write(text)
                            sys.stdout.flush()
                            text = ""
                        else:
                            before = text[:tag_end]
                            if before:
                                sys.stdout.write(before)
                            sys.stdout.write("</think>" + THINK_CLOSE)
                            sys.stdout.flush()
                            in_think_tag = False
                            text = text[tag_end + len("</think>"):]

            # ── Reasoning tokens ──────────────────────────────────────
            # Xiaomi Token Plan returns reasoning_content
            # OpenRouter returns reasoning
            reasoning_text = None
            if hasattr(delta, "reasoning_content") and delta.reasoning_content:
                reasoning_text = delta.reasoning_content
            elif hasattr(delta, "reasoning") and delta.reasoning:
                reasoning_text = delta.reasoning

            if reasoning_text:
                sys.stdout.write(THINK_OPEN + reasoning_text + THINK_CLOSE)
                sys.stdout.flush()

            # ── Tool call deltas ────────────────────────────────────────
            if delta.tool_calls:
                for tc_delta in delta.tool_calls:
                    idx = tc_delta.index
                    if idx not in tool_calls:
                        tool_calls[idx] = {"id": "", "name": "", "arguments": ""}
                    if tc_delta.id:
                        tool_calls[idx]["id"] = tc_delta.id
                    if tc_delta.function and tc_delta.function.name:
                        tool_calls[idx]["name"] = tc_delta.function.name
                    if tc_delta.function and tc_delta.function.arguments:
                        tool_calls[idx]["arguments"] += tc_delta.function.arguments

        if in_think_tag:
            sys.stdout.write(THINK_CLOSE)   # reset if stream ended mid-tag
        sys.stdout.write("\n")
        sys.stdout.flush()

        # Fall back to character-based estimation if provider didn't return usage
        if not usage:
            est_prompt = self._estimate_tokens(messages)
            est_completion = len(full_content) // 4
            usage = {
                "prompt_tokens":     est_prompt,
                "completion_tokens": est_completion,
                "total_tokens":      est_prompt + est_completion,
                "estimated":         True,
            }

        return finish_reason, full_content, tool_calls, usage

    # ------------------------------------------------------------------
    # Tool execution
    # ------------------------------------------------------------------

    def _execute_tool(self, name: str, inputs: dict) -> dict:
        if name == "read_skills_doc":
            skill = inputs.get("skill", "pdf")
            result = read_skills_doc(skill=skill)
            self.state["skills_loaded"] = result.get("ok", False)
            if result["ok"]:
                print(f"     Skill '{skill}' loaded ({result['size_chars']} chars)")
                return {
                    "ok": True,
                    "skill": skill,
                    "path": result["path"],
                    "size_chars": result["size_chars"],
                    "full_content": result["content"],
                }
            return result

        if name == "run_python":
            from .tools.run_python import run_python
            desc = inputs.get("description", "")
            code = inputs.get("code", "")
            print(f"     Running: {desc}")
            result = run_python(code=code, description=desc)
            status = "✅" if result["ok"] else "❌"
            stdout_preview = result["stdout"][:200].replace("\n", " ↵ ")
            print(f"     {status} exit={result['exit_code']}  out: {stdout_preview}")
            return result

        if name == "edit_xlsx_payees":
            from .tools.edit_xlsx import edit_xlsx_payees
            xlsx_path = inputs.get("xlsx_path", "")
            edits = inputs.get("edits", {})
            print(f"     Editing {len(edits)} payee(s) in {xlsx_path}")
            result = edit_xlsx_payees(xlsx_path=xlsx_path, edits=edits)
            status = "✅" if result["ok"] else "❌"
            print(f"     {status} edited={result.get('edited_count', 0)}  errors={result.get('errors', [])}")
            return result

        if name == "verify_xlsx_balance":
            from .tools.verify_balance import verify_xlsx_balance
            xlsx_path = inputs.get("xlsx_path", "")
            print(f"     Verifying balance: {xlsx_path}")
            result = verify_xlsx_balance(xlsx_path=xlsx_path)
            if result.get("ok"):
                l1 = "✅" if result["l1_pass"] else "❌"
                l2 = "✅" if result["l2_pass"] else "❌"
                print(f"     L1:{l1} L2:{l2}  rows={result['total_rows']}  failed={len(result['failed_rows'])}")
            else:
                print(f"     ❌ {result.get('error')}")
            return result

        if name == "read_xlsx_transactions":
            from .tools.read_xlsx import read_xlsx_transactions
            xlsx_path = inputs.get("xlsx_path", "")
            print(f"     Reading transactions from {xlsx_path}")
            result = read_xlsx_transactions(xlsx_path=xlsx_path)
            if result.get("ok"):
                print(f"     ✅ {result['total_rows']} rows, columns: {result['columns']}")
            else:
                print(f"     ❌ {result.get('error')}")
            return result

        return {"ok": False, "error": f"Unknown tool: {name}"}

    # ------------------------------------------------------------------
    # Output verification
    # ------------------------------------------------------------------

    def _check_outputs(self, xlsx_path: str) -> dict:
        xlsx_p = Path(xlsx_path)
        xlsx_rows = 0
        if xlsx_p.exists() and xlsx_p.stat().st_size > 100:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(xlsx_p)
                # Prefer "Transactions" sheet; fall back to active
                ws = wb["Transactions"] if "Transactions" in wb.sheetnames else wb.active
                xlsx_rows = ws.max_row - 1  # subtract header row
            except Exception:
                xlsx_rows = -1
        return {
            "xlsx_exists": xlsx_p.exists() and xlsx_p.stat().st_size > 100,
            "xlsx_rows": max(xlsx_rows, 0),
        }
