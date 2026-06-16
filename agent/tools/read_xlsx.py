"""
read_xlsx_transactions — Fast structured read of the Transactions sheet.
Returns all rows as a list of dicts keyed by column header.
"""

from pathlib import Path


def read_xlsx_transactions(xlsx_path: str, max_rows: int = 2000) -> dict:
    """
    Read all transactions from the Transactions sheet.

    Returns:
        {
            ok,
            columns: [str],           # column header names in order
            rows: [                   # each row as a dict
                {
                    _row: int,        # openpyxl row number (2 = first data row)
                    page, date, description, payee, credit, debit, balance
                }
            ],
            total_rows: int
        }
    """
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl not available", "rows": []}

    xlsx_p = Path(xlsx_path)
    if not xlsx_p.exists():
        return {"ok": False, "error": f"File not found: {xlsx_path}", "rows": []}

    try:
        wb = openpyxl.load_workbook(xlsx_p, data_only=True, read_only=True)
    except Exception as exc:
        return {"ok": False, "error": f"Cannot open workbook: {exc}", "rows": []}

    if "Transactions" not in wb.sheetnames:
        return {"ok": False, "error": "Sheet 'Transactions' not found", "rows": []}

    ws = wb["Transactions"]

    headers = []
    rows = []
    row_num = 1

    for xlsx_row in ws.iter_rows(values_only=False):
        if row_num == 1:
            # Header row
            headers = [
                str(cell.value).strip() if cell.value is not None else f"col{i}"
                for i, cell in enumerate(xlsx_row)
            ]
            row_num += 1
            continue

        if row_num - 1 > max_rows:
            break

        values = [cell.value for cell in xlsx_row]

        # Skip completely blank rows
        if all(v is None for v in values):
            row_num += 1
            continue

        row_dict = {"_row": row_num}
        for i, header in enumerate(headers):
            row_dict[header.lower()] = values[i] if i < len(values) else None

        rows.append(row_dict)
        row_num += 1

    wb.close()

    return {
        "ok": True,
        "columns": headers,
        "rows": rows,
        "total_rows": len(rows),
    }
