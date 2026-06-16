"""
verify_xlsx_balance — Structured balance verification for the Transactions sheet.
Returns L1 (aggregate) and L2 (row-by-row) pass/fail with exact failing rows.
"""

from pathlib import Path


def _to_float(v) -> float:
    """Convert cell value to float, treating None / empty string as 0.0."""
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return 0.0


def verify_xlsx_balance(xlsx_path: str) -> dict:
    """
    Verify balance integrity of the Transactions sheet.

    L1 (aggregate): opening_balance + sum(credits) - sum(debits) == closing_balance (±0.01)
    L2 (row-by-row): for every consecutive pair, prev_balance + credit - debit == current_balance (±0.01)

    Returns:
        {
            ok,
            l1_pass, l2_pass,
            opening_balance, closing_balance,
            sum_credits, sum_debits, computed_closing,
            l1_diff,                          # computed_closing - closing_balance
            failed_rows: [                    # L2 failures only
                {row, date, description, credit, debit, balance, expected_balance, diff}
            ]
        }
    """
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl not available"}

    xlsx_p = Path(xlsx_path)
    if not xlsx_p.exists():
        return {"ok": False, "error": f"File not found: {xlsx_path}"}

    try:
        wb = openpyxl.load_workbook(xlsx_p, data_only=True)
    except Exception as exc:
        return {"ok": False, "error": f"Cannot open workbook: {exc}"}

    if "Transactions" not in wb.sheetnames:
        return {"ok": False, "error": "Sheet 'Transactions' not found"}

    ws = wb["Transactions"]

    # Discover column positions from header row
    col_map = {}
    for cell in ws[1]:
        if cell.value:
            col_name = str(cell.value).strip().lower()
            col_map[col_name] = cell.column - 1  # 0-based index
            # Also map variations: "Credit (RM)" → "credit"
            col_name_base = col_name.split("(")[0].strip()
            if col_name_base and col_name_base != col_name:
                col_map[col_name_base] = cell.column - 1

    required = {"credit", "debit", "balance"}
    missing = required - set(col_map)
    if missing:
        return {"ok": False, "error": f"Missing columns: {missing}"}

    credit_idx  = col_map["credit"]
    debit_idx   = col_map["debit"]
    balance_idx = col_map["balance"]
    date_idx    = col_map.get("date", -1)
    desc_idx    = col_map.get("description", -1)

    # Read all data rows
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue  # skip blank rows
        rows.append(r)

    if not rows:
        return {"ok": False, "error": "No data rows found"}

    # Compute sums
    sum_credits = sum(_to_float(r[credit_idx]) for r in rows)
    sum_debits  = sum(_to_float(r[debit_idx])  for r in rows)

    first_row  = rows[0]
    last_row   = rows[-1]
    closing_balance = _to_float(last_row[balance_idx])

    # Opening balance: reverse-compute from first row
    opening_balance = (
        _to_float(first_row[balance_idx])
        - _to_float(first_row[credit_idx])
        + _to_float(first_row[debit_idx])
    )

    computed_closing = round(opening_balance + sum_credits - sum_debits, 2)
    l1_diff  = round(computed_closing - closing_balance, 2)
    l1_pass  = abs(l1_diff) <= 0.01

    # L2: row-by-row chain check
    failed_rows = []
    prev_balance = opening_balance

    for i, r in enumerate(rows):
        credit  = _to_float(r[credit_idx])
        debit   = _to_float(r[debit_idx])
        balance = _to_float(r[balance_idx])
        expected = round(prev_balance + credit - debit, 2)
        diff = round(balance - expected, 2)

        if abs(diff) > 0.01:
            failed_rows.append({
                "row": i + 2,  # openpyxl row number (header=1)
                "date": r[date_idx] if date_idx >= 0 else None,
                "description": (str(r[desc_idx])[:80] if desc_idx >= 0 else None),
                "credit": credit,
                "debit": debit,
                "balance": balance,
                "expected_balance": expected,
                "diff": diff,
            })

        prev_balance = balance

    l2_pass = len(failed_rows) == 0

    return {
        "ok": True,
        "l1_pass": l1_pass,
        "l2_pass": l2_pass,
        "opening_balance": round(opening_balance, 2),
        "closing_balance": round(closing_balance, 2),
        "sum_credits": round(sum_credits, 2),
        "sum_debits": round(sum_debits, 2),
        "computed_closing": computed_closing,
        "l1_diff": l1_diff,
        "total_rows": len(rows),
        "failed_rows": failed_rows,
    }
