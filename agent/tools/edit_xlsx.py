"""
edit_xlsx_payees — Precise payee editing without full re-extraction.
Takes a dict of {openpyxl_row_number: new_payee} and writes only those cells.
Row numbering: row 1 = header, row 2 = first data row.
"""

from pathlib import Path


def edit_xlsx_payees(xlsx_path: str, edits: dict) -> dict:
    """
    Precisely edit Payee cells in the Transactions sheet.

    edits: mapping of row_number (int or str) → new_payee (str)
           Row numbers are openpyxl 1-based where row 1 = header, row 2 = first data row.

    Returns {ok, edited_count, errors, xlsx_path}
    """
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl not available", "edited_count": 0, "errors": []}

    xlsx_p = Path(xlsx_path)
    if not xlsx_p.exists():
        return {"ok": False, "error": f"File not found: {xlsx_path}", "edited_count": 0, "errors": []}

    try:
        wb = openpyxl.load_workbook(xlsx_p)
    except Exception as exc:
        return {"ok": False, "error": f"Cannot open workbook: {exc}", "edited_count": 0, "errors": []}

    if "Transactions" not in wb.sheetnames:
        return {"ok": False, "error": "Sheet 'Transactions' not found", "edited_count": 0, "errors": []}

    ws = wb["Transactions"]

    # Find Payee column by header name (row 1)
    payee_col = None
    for cell in ws[1]:
        if cell.value and str(cell.value).strip().lower() == "payee":
            payee_col = cell.column
            break

    if payee_col is None:
        return {"ok": False, "error": "Payee column not found in header row", "edited_count": 0, "errors": []}

    max_row = ws.max_row
    edited = 0
    errors = []

    for row_key, new_payee in edits.items():
        try:
            row_num = int(row_key)
        except (ValueError, TypeError):
            errors.append(f"Invalid row key: {row_key!r}")
            continue

        if row_num < 2:
            errors.append(f"Row {row_num} is the header — skipped")
            continue
        if row_num > max_row:
            errors.append(f"Row {row_num} exceeds sheet max_row={max_row}")
            continue

        ws.cell(row=row_num, column=payee_col).value = str(new_payee).strip()
        edited += 1

    try:
        wb.save(xlsx_p)
    except Exception as exc:
        return {"ok": False, "error": f"Failed to save workbook: {exc}", "edited_count": edited, "errors": errors}

    return {
        "ok": True,
        "edited_count": edited,
        "errors": errors,
        "xlsx_path": xlsx_path,
    }
